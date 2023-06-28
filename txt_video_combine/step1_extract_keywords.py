import os
import openpyxl
from tqdm import tqdm
from docx import Document
from segtok.segmenter import split_single
from concurrent.futures import ThreadPoolExecutor, as_completed
import openai
import time

openai.api_key = "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"

def request_with_retry(messages, max_tokens=500, max_requests=90, cooldown_seconds=60):
    while True:
        try:
            response = openai.ChatCompletion.create(
                model="gpt-3.5-turbo",
                messages=messages,
                max_tokens=max_tokens,
                n=1,
                stop=None
            )
            return response['choices'][0]['message']['content'].strip()
        except openai.error.RateLimitError:
            print("超过速率限制。正在等待冷却时间...")
            time.sleep(cooldown_seconds)
        except Exception as e:
            print(f"发生错误：{str(e)}")
            time.sleep(10)


def translate_text(text):
    messages = [
        {"role": "system", "content": "你是一个有帮助的助手。"},
        {"role": "user", "content": f"将以下中文文本翻译为英文：{text}"},
    ]
    return request_with_retry(messages)


def read_docx(file_path):
    document = Document(file_path)
    paragraphs = [paragraph.text for paragraph in document.paragraphs if paragraph.text.strip()]
    if not paragraphs:
        raise ValueError("未能读取到有效的文本内容")
    return paragraphs


def write_to_excel(sentences, workbook):
    sheet = workbook.active
    for idx, sentence in enumerate(sentences, 1):
        sheet.cell(row=idx, column=1, value=sentence)


def split_sentences(paragraph):
    return split_single(paragraph)


def process_text_sentences(workbook, input_file_path, output_file_path):
    try:
        paragraphs = read_docx(input_file_path)
    except ValueError as e:
        print(f"发生错误：{str(e)}")
        return

    sentences = [sentence for paragraph in paragraphs for sentence in split_sentences(paragraph)]
    write_to_excel(sentences, workbook)

    sheet = workbook.active
    max_workers = min(len(sentences), 5)
    row_idx = 1

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = []
        for sentence in sentences:
            stripped_sentence = sentence.strip()
            if not stripped_sentence:
                sheet.cell(row=row_idx, column=2, value="")
                row_idx += 1
                continue

            futures.append(executor.submit(translate_text, stripped_sentence))
            row_idx += 1

        for future in tqdm(as_completed(futures), total=len(futures), desc='中译英翻译中'):
            index = futures.index(future)
            translated_text = future.result()
            sheet.cell(row=index + 1, column=2, value=translated_text)

    workbook.save(output_file_path)

def main():
    input_file_path = 'F:/txt_to_video/input.docx'
    output_file_path = 'F:/txt_to_video/prompt.txt/prompt.txt.xlsx'
    workbook = openpyxl.Workbook()
    process_text_sentences(workbook, input_file_path, output_file_path)

if __name__ == "__main__":
    main()