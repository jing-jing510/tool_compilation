# openpyxl 是用于读取和写入Excel 2010 xlsx / xlsm / xltx / xltm 文件的Python 库  不支持csv格式
import os
import csv
import chardet
import openpyxl

# 读取novel/分镜设置.csv 文件并打印出来
def read_csv_file():
    # with open(file_name, 'r') as f:
    #     lines = f.readlines()
    #     for line in lines:
    #         print(line)
    # with open(,'rb') as f:
    #     raw_data =f.read()
    #     detected_encoding = chardet.detect(raw_data)['encoding']
    #     print(detected_encoding)
    # fenjing_file = os.path.join('novel',file_name)
    # print(fenjing_file)
    # wb = openpyxl.load_workbook(fenjing_file)

    # # openpyxl.utils.exceptions.InvalidFileException: openpyxl does not support .csv file format
    wb = openpyxl.load_workbook('novel/分镜设置.xlsx')
    ws = wb.active
    c = ws['B2']

    print(c.value)

if __name__ == '__main__':
    read_csv_file()
    print(os.path.abspath(__file__))
    print(os.path.dirname(os.path.abspath(__file__)))
    print(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    # current_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # print(csv.reader('novel/分镜设置.csv'))
    with open('novel/分镜设置.csv','rb') as f:
        raw_data = f.read()
        detected_encoding = chardet.detect(raw_data)['encoding']
        print(detected_encoding)
    with open('novel/分镜设置.csv','r',encoding=detected_encoding,errors='ignore') as f:

        reader = csv.reader(f)
        for row in reader:
            print(row)