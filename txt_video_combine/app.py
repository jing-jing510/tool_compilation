# auth ： xiaokou
# date ： 2023/6/28 11:00
import tkinter as tk
from tkinter import ttk

import openpyxl

import activation
import openai_text
import threading

# 创建主窗口
root = tk.Tk()
root.title("文本转换")
root.geometry("1920x1080")

# 第一行:软件激活码
label_activation = tk.Label(root, text="激活码")
label_activation.grid(row=0, column=0)

entry_activation = tk.Entry(root)
entry_activation.grid(row=0, column=1)

button_activation = tk.Button(root, text="激活软件",
                              command=lambda: activation.check_activation_code(entry_activation.get()))
button_activation.grid(row=0, column=2)

# 第一行:GPT端口
label_port = tk.Label(root, text="GPT端口:")
label_port.grid(row=0, column=3)

port_list = ['gpt-3.5-turbo', 'gpt-4.0-turbo']
port_variable = tk.StringVar(root)
port_variable.set(port_list[0])  # 默认值

optionmenu_port = tk.OptionMenu(root, port_variable, *port_list)
optionmenu_port.grid(row=0, column=4)
# 第一行：GPT key
entry_key = tk.Entry(root)
entry_key.grid(row=0, column=6)

button_set_gpt_key = tk.Button(root, text="第一步:GPT Key",
                               command=lambda: openai_text.set_gpt_key(entry_key.get()))
button_set_gpt_key.grid(row=0, column=7)


# 创建一个按钮,执行openai_text.ai_fenjing_array() 操作,将其数组lines 放到表格当中
def insert_to_table():
    lines = openai_text.ai_fenjing_array()
    # print(lines)
    for i, line in enumerate(lines):
        print(f"正在插入的列:{i}:{line}")
        table.insert('', 'end', values=(line, '', '', '', ''), iid=str(i + 1))


button_ai_fenjing_array = tk.Button(root, text="第二步:原文",
                                    command=insert_to_table)
button_ai_fenjing_array.grid(row=1, column=0)


# 进行分镜操作
# def process_all_fenjing():
#     # 循环遍历表格中的所有行
#     for item in table.get_children():
#         original_text = table.item(item)['values'][0]  # 获取这一行"原文"列的内容
#         result = openai_text.ai_fenjing(original_text)  # 调用ai_fenjing函数
#
#         # 更新表格中的对应行的“分镜”列
#         table.set(item, "Fenjing", result)
#
#         root.update_idletasks()  # 刷新界面
#
#
# def start_process_all_fenjing():
#     threading.Thread(target=process_all_fenjing).start()
# def process_all_fenjing():
#     items = table.get_children()
#     results = []
#
#     for item in items:
#         original_text = table.item(item)['values'][0]  # 获取这一行"原文"列的内容
#         result = openai_text.ai_fenjing(original_text)  # 调用ai_fenjing函数
#         results.append(result)
#     # 将结果保存到Excel文件
#     save_fenjing_to_excel(results)
#
#
# def save_fenjing_to_excel(results):
#     # 打开现有的 Excel 文件（如果存在），否则创建一个新的
#     try:
#         workbook = openpyxl.load_workbook("output/fenjing.xlsx")
#     except FileNotFoundError:
#         workbook = openpyxl.Workbook()
#
#     # 选择第一个工作表
#     sheet = workbook.active
#
#     # 如果表格不为空，则删除所有行
#     if sheet.max_row > 1:
#         sheet.delete_rows(2, sheet.max_row)
#
#     # 将结果插入表格
#     for i, result in enumerate(results, start=2):  # 从第二行开始插入数据
#         sheet.cell(row=i, column=1, value=result)
#
#     # 保存 Excel 文件
#     workbook.save("output/fenjing.xlsx")
#     workbook.close()
#
# def start_process_all_fenjing():
#     threading.Thread(target=process_all_fenjing).start()

def process_all_fenjing():
    items = table.get_children()
    total_items = len(items)
    progress["maximum"] = total_items  # 设置进度条的最大值为表格中的行数
    for i, item in enumerate(items, start=1):
        original_text = table.item(item)['values'][0]
        result = openai_text.ai_fenjing(original_text)
        result = result.replace("镜头语言:","")
        # 更新表格中的对应行的“分镜”列
        table.set(item, "Fenjing", result)

        # 更新进度条和进度标签
        root.after(0, progress.configure, {"value": i})  # 使用root.after将进度条更新操作交给Tkinter主线程
        root.after(0, progress_label.configure, {"text": f"Processing: {i}/{total_items}"})
        root.update_idletasks()  # 更新窗口，显示最新的GUI状态

def start_process_all_fenjing():
    threading.Thread(target=process_all_fenjing).start()

# 创建进度条和进度标签
progress = ttk.Progressbar(root, mode="determinate", maximum=100)  # 初始化进度条的最大值为100
progress_label = tk.Label(root, text="Processing: 0/0")

# 将进度条和进度标签添加到界面
progress_label.grid(row=1, column=2)
progress.grid(row=1, column=3)

button_fenjing = tk.Button(root, text="第三步:分镜", command=start_process_all_fenjing)  # 创建按钮，并绑定process_fenjing函数
button_fenjing.grid(row=1, column=1)  # 添加按钮到界面

def process_all_prompt():
    items = table.get_children()
    total_items = len(items)
    progress["maximum"] = total_items  # 设置进度条的最大值为表格中的行数
    for i, item in enumerate(items, start=1):
        original_text = table.item(item)['values'][1]
        result = openai_text.ai_prompt(original_text)

        # 更新表格中的对应行的“分镜”列
        table.set(item, "Positive", result)

        # 更新进度条和进度标签
        root.after(0, progress.configure, {"value": i})  # 使用root.after将进度条更新操作交给Tkinter主线程
        root.after(0, progress_label.configure, {"text": f"Processing: {i}/{total_items}"})
        root.update_idletasks()  # 更新窗口，显示最新的GUI状态

def start_process_all_prompt():
    threading.Thread(target=process_all_prompt).start()

# 创建进度条和进度标签
progress1 = ttk.Progressbar(root, mode="determinate", maximum=100)  # 初始化进度条的最大值为100
progress_label1 = tk.Label(root, text="Processing: 0/0")

# 将进度条和进度标签添加到界面
progress_label1.grid(row=1, column=5)
progress1.grid(row=1, column=6)

button_fenjing = tk.Button(root, text="第四部:提示词", command=start_process_all_prompt)  # 创建按钮，并绑定process_prompt函数
button_fenjing.grid(row=1, column=4)  # 添加按钮到界面

# 创建一个表格 第一行 原文 第二行 分镜 第三行 正向关键词 第四行 反向关键词 第五行 图片
# 创建表格
table = ttk.Treeview(root, columns=("Original", "Fenjing", "Positive", "Negative", "Image", "ReDraw"), show='headings',
                     xscrollcommand=True)
# # 创建滚动条
# scrollbar = ttk.Scrollbar(root, orient="horizontal", command=table.xview)
# scrollbar.grid(row=2, column=0, sticky="we", columnspan=5)  # 根据需要更改row、column值
# table.configure(xscrollcommand=scrollbar.set)
# 设置表头
table.heading("Original", text="原文")
table.heading("Fenjing", text="分镜")
table.heading("Positive", text="正向关键词")
table.heading("Negative", text="反向关键词")
table.heading("Image", text="图片")
table.heading("ReDraw", text="重绘")


# 设置列的宽度
def resize_columns(event):
    # 获取当前窗口的宽度
    width = root.winfo_width()
    # print(width)
    num_columns = 6  # 列的数量
    new_width = width // num_columns - 1  # 计算新的列宽度
    table.column("Original", width=new_width)
    table.column("Fenjing", width=new_width)
    table.column("Positive", width=new_width)
    table.column("Negative", width=new_width)
    table.column("Image", width=new_width)
    table.column("ReDraw", width=new_width)


# 当窗口大小发生改变时，调用resize_columns函数
root.bind("<Configure>", resize_columns)


# 添加表格到界面
table.grid(row=2, column=0, columnspan=100, sticky="nsew")
# 配置表格在窗口大小改变时自动扩展
root.grid_rowconfigure(2, weight=1)

# 运行主循环，显示窗口
root.mainloop()
